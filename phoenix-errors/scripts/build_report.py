#!/usr/bin/env python3
"""
Render a Phoenix error report from pulled span data.

Inputs (JSON, produced by Claude in Phase B):
  --spans     spans.json     flat array of error-span records (required)
  --analysis  analysis.json  narrative layer: families, recommendations, etc. (optional)

Outputs:
  --out <prefix>   writes <prefix>.xlsx and/or <prefix>.html
  --format both|xlsx|html

The script is intentionally "dumb": it computes counts and aggregations and applies
styling, but every piece of domain judgment (which families exist, their root cause,
severity, recommendations) comes from analysis.json. See references/data-pull.md for
the input schemas.

Dependency: openpyxl (for xlsx). HTML output needs no third-party deps.
"""
import argparse
import html
import json
import sys
from collections import defaultdict, OrderedDict, Counter
from datetime import datetime

LOOP_THRESHOLD = 5  # a (session, message-id) seen >= this many times is flagged as a redelivery loop

# Severity → fill colors (shared by xlsx + html)
SEV_COLORS = {
    "P0": "F8D7DA",  # red
    "P1": "FFE8CC",  # orange
    "P2": "FFF3CD",  # yellow
}
DEFAULT_FILL = "EEEEEE"

SPAN_FIELDS = [
    "project", "span_id", "trace_id", "start_time", "end_time", "name",
    "span_kind", "status", "user_id", "session_id", "model",
    "exception_type", "error_family", "exception_message",
]
SPAN_HEADERS = [
    "Project", "Span ID", "Trace ID", "Start (UTC)", "End (UTC)", "Span Name",
    "Kind", "Status", "User ID", "Session ID", "Model",
    "Exception Type", "Error Family", "Exception Message",
]


# ---------------- helpers ----------------

def heuristic_family(exc_type, exc_msg):
    """Fallback family label when Claude didn't supply error_family."""
    t, m = (exc_type or ""), (exc_msg or "")
    tl, ml = t.lower(), m.lower()
    if "ratelimit" in tl or "429" in ml or "rate limit" in ml:
        return "Rate limit (429)"
    if "badrequest" in tl or "structured output" in ml or "schema" in ml:
        return "Bad request / schema failure"
    if "timeout" in tl or "timed out" in ml:
        return "Timeout"
    if "auth" in tl or "unauthorized" in ml or "401" in ml or "403" in ml:
        return "Auth failure"
    if t:
        return t
    return "Other"


def _to_int(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str) and v.strip().isdigit():
        return int(v)
    return None


def _duration_s(rec, raw):
    """Seconds between start and end. Prefer explicit duration_s, else compute from timestamps."""
    if raw.get("duration_s") not in (None, ""):
        try:
            return float(raw["duration_s"])
        except (TypeError, ValueError):
            pass
    s, e = rec.get("start_time", ""), rec.get("end_time", "")
    try:
        return (datetime.fromisoformat(e) - datetime.fromisoformat(s)).total_seconds()
    except (TypeError, ValueError):
        return None


def load_spans(path, default_status="ERROR"):
    with open(path) as f:
        rows = json.load(f)
    norm = []
    for r in rows:
        rec = {k: (r.get(k) or "") for k in SPAN_FIELDS}
        if not rec["error_family"] and default_status == "ERROR":
            rec["error_family"] = heuristic_family(rec["exception_type"], rec["exception_message"])
        if not rec["status"]:
            rec["status"] = default_status
        # diagnostics-only fields (not shown in the All Errors sheet)
        rec["duration_s"] = _duration_s(rec, r)
        rec["prompt_tokens"] = _to_int(r.get("prompt_tokens"))
        rec["completion_tokens"] = _to_int(r.get("completion_tokens"))
        rec["total_tokens"] = _to_int(r.get("total_tokens"))
        rec["mid"] = r.get("mid") or ""
        norm.append(rec)
    return norm


def load_analysis(path):
    if not path:
        return {}
    with open(path) as f:
        return json.load(f)


def severity_for_family(fam, analysis):
    for f in analysis.get("families", []):
        if f.get("family") == fam:
            return f.get("severity", "P2")
    return "P2"


def family_meta(fam, analysis):
    for f in analysis.get("families", []):
        if f.get("family") == fam:
            return f.get("root_cause", ""), f.get("action", "")
    return "", ""


def aggregate(spans, analysis):
    total = len(spans)

    # By family
    fam_counts = OrderedDict()
    for s in spans:
        fam_counts[s["error_family"]] = fam_counts.get(s["error_family"], 0) + 1
    # order families by analysis order first, then by count desc
    ordered_fams = [f["family"] for f in analysis.get("families", []) if f["family"] in fam_counts]
    for fam in sorted(fam_counts, key=lambda k: -fam_counts[k]):
        if fam not in ordered_fams:
            ordered_fams.append(fam)

    # By project — seed from analysis.projects (so clean projects appear), then add any seen
    proj_seen = OrderedDict()
    for p in analysis.get("projects", []):
        proj_seen[p["name"]] = {"errors": 0, "traces": set(), "users": set(), "note": p.get("note", "")}
    for s in spans:
        d = proj_seen.setdefault(s["project"], {"errors": 0, "traces": set(), "users": set(), "note": ""})
        d["errors"] += 1
        if s["trace_id"]:
            d["traces"].add(s["trace_id"])
        if s["user_id"]:
            d["users"].add(s["user_id"])

    # By trace
    traces = defaultdict(lambda: {"project": "", "spans": [], "user": "", "fams": set()})
    for s in spans:
        td = traces[s["trace_id"]]
        td["project"] = s["project"]
        td["user"] = s["user_id"]
        td["spans"].append((s["start_time"], s["end_time"]))
        td["fams"].add(s["error_family"])

    # By user
    users = defaultdict(lambda: {"project": "", "count": 0, "traces": set()})
    for s in spans:
        key = (s["user_id"], s["project"])
        users[key]["count"] += 1
        if s["trace_id"]:
            users[key]["traces"].add(s["trace_id"])

    distinct_traces = len({s["trace_id"] for s in spans if s["trace_id"]})
    distinct_users = len({s["user_id"] for s in spans if s["user_id"]})
    times = sorted([s["start_time"] for s in spans if s["start_time"]])
    window = f"{times[0][11:16]} – {times[-1][11:16]}" if times else "—"

    return {
        "total": total,
        "fam_counts": fam_counts,
        "ordered_fams": ordered_fams,
        "projects": proj_seen,
        "traces": traces,
        "users": users,
        "distinct_traces": distinct_traces,
        "distinct_users": distinct_users,
        "window": window,
    }


# ---------------- diagnostics ----------------

def _percentile(vals, p):
    xs = sorted(v for v in vals if v is not None)
    if not xs:
        return None
    k = (len(xs) - 1) * p / 100.0
    lo = int(k)
    hi = min(lo + 1, len(xs) - 1)
    return xs[lo] + (xs[hi] - xs[lo]) * (k - lo)


def _stats(vals):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    return {"n": len(vals), "avg": sum(vals) / len(vals),
            "p50": _percentile(vals, 50), "p95": _percentile(vals, 95)}


def diagnostics(spans, ok_spans, analysis):
    """Compute the four diagnostic views. Pure structural math; no domain judgment."""
    ok_spans = ok_spans or []

    # 1. Latency by project × span-kind, for ERROR (time-to-fail) and OK (healthy)
    def lat_table(rows):
        buckets = defaultdict(list)
        for r in rows:
            buckets[(r["project"], r.get("span_kind") or "?")].append(r.get("duration_s"))
        return {k: _stats(v) for k, v in sorted(buckets.items())}

    latency_error = lat_table(spans)
    latency_ok = lat_table(ok_spans)

    # 2. Token usage by model (from any span that carries token counts — mostly healthy LLM spans)
    tok = defaultdict(lambda: {"prompt": [], "completion": [], "total": []})
    for r in spans + ok_spans:
        if r.get("prompt_tokens") or r.get("completion_tokens") or r.get("total_tokens"):
            m = r.get("model") or "?"
            for key, fld in (("prompt", "prompt_tokens"), ("completion", "completion_tokens"), ("total", "total_tokens")):
                if r.get(fld) is not None:
                    tok[m][key].append(r[fld])
    tokens = {m: {k: _stats(v) for k, v in d.items()} for m, d in tok.items()}

    # 3. Concentration + redelivery-loop detection
    user_err = Counter(r["user_id"] for r in spans if r["user_id"])
    sess_err = Counter(r["session_id"] for r in spans if r["session_id"])
    total_err = len(spans)
    top_user_pct = (user_err.most_common(1)[0][1] / total_err * 100) if (user_err and total_err) else 0
    mid_counts = Counter((r["session_id"], r["mid"]) for r in spans if r.get("mid"))
    loops = [{"session": s, "mid": m, "count": c}
             for (s, m), c in mid_counts.most_common(15) if c >= LOOP_THRESHOLD]

    # 4. Errors over time (hourly buckets, UTC)
    hours = Counter(r["start_time"][:13] for r in spans if r["start_time"])
    over_time = sorted(hours.items())

    # 5. Error rate per project — true rate if analysis.totals given, else sampled from OK pull, else N/A
    totals = analysis.get("totals", {})
    proj_names = set(r["project"] for r in spans) | set(totals) | {p["name"] for p in analysis.get("projects", [])}
    error_rate = {}
    for p in sorted(proj_names):
        err = sum(1 for r in spans if r["project"] == p)
        if p in totals:
            tot = totals[p].get("total") or (totals[p].get("errors", 0) + totals[p].get("ok", 0))
            error_rate[p] = {"errors": err, "total": tot,
                             "pct": (err / tot * 100 if tot else None), "basis": "reported totals"}
        elif ok_spans:
            okc = sum(1 for r in ok_spans if r["project"] == p)
            tot = err + okc
            error_rate[p] = {"errors": err, "total": tot,
                             "pct": (err / tot * 100 if tot else None), "basis": "SAMPLED (err+ok sample)"}
        else:
            error_rate[p] = {"errors": err, "total": None, "pct": None, "basis": "errors-only"}

    return {
        "latency_error": latency_error, "latency_ok": latency_ok,
        "tokens": tokens,
        "top_users": user_err.most_common(10), "top_sessions": sess_err.most_common(10),
        "top_user_pct": top_user_pct,
        "loops": loops,
        "over_time": over_time,
        "error_rate": error_rate,
        "has_ok": bool(ok_spans),
    }


# ---------------- XLSX ----------------

def build_xlsx(spans, analysis, agg, diag, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", start_color="1F4E78")
    TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
    SUB_FONT = Font(name="Arial", bold=True, size=11, color="1F4E78")
    BODY = Font(name="Arial", size=10)
    MONO = Font(name="Consolas", size=9)
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="top")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def fill_for(sev):
        return PatternFill("solid", start_color=SEV_COLORS.get(sev, DEFAULT_FILL))

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

    def widths(ws, ws_widths):
        for i, w in enumerate(ws_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # ----- Summary -----
    ws = wb.active
    ws.title = "Summary"
    title = analysis.get("report_title", "Phoenix Error Report")
    ws["A1"] = title; ws["A1"].font = TITLE_FONT; ws.merge_cells("A1:E1")
    ws["A2"] = f"Date: {analysis.get('date', '')}    Window: {analysis.get('window', '')}"
    ws["A2"].font = BODY; ws.merge_cells("A2:E2")
    ws["A3"] = f"Source: Arize Phoenix ({analysis.get('source', '')})"
    ws["A3"].font = BODY; ws.merge_cells("A3:E3")

    ws["A5"] = "Totals"; ws["A5"].font = SUB_FONT
    totals = [
        ("Total error spans", agg["total"]),
        ("Distinct traces", agg["distinct_traces"]),
        ("Distinct users affected", agg["distinct_users"]),
        ("Time window (UTC)", agg["window"]),
    ]
    for i, (k, v) in enumerate(totals):
        ws.cell(row=6 + i, column=1, value=k).font = BODY
        ws.cell(row=6 + i, column=2, value=v).font = BODY

    # By Project
    ws["A11"] = "By Project"; ws["A11"].font = SUB_FONT
    for i, h in enumerate(["Project", "Errors", "Distinct Traces", "Distinct Users", "Note"], 1):
        ws.cell(row=12, column=i, value=h)
    style_header(ws, 12, 5)
    r = 13
    for name, d in agg["projects"].items():
        vals = [name, d["errors"], len(d["traces"]), len(d["users"]), d.get("note", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.font = BODY; cell.border = BORDER
        r += 1

    # By Error Family
    fam_start = r + 1
    ws.cell(row=fam_start, column=1, value="By Error Family").font = SUB_FONT
    for i, h in enumerate(["Error Family", "Count", "% of Total", "Root Cause", "Recommended Action"], 1):
        ws.cell(row=fam_start + 1, column=i, value=h)
    style_header(ws, fam_start + 1, 5)
    rr = fam_start + 2
    for fam in agg["ordered_fams"]:
        cnt = agg["fam_counts"][fam]
        pct = (cnt / agg["total"]) if agg["total"] else 0
        rc, action = family_meta(fam, analysis)
        sev = severity_for_family(fam, analysis)
        vals = [fam, cnt, pct, rc, action]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = BODY; cell.alignment = WRAP; cell.border = BORDER
            cell.fill = fill_for(sev)
        ws.cell(row=rr, column=3).number_format = "0.0%"
        ws.row_dimensions[rr].height = 42
        rr += 1
    widths(ws, [42, 14, 12, 50, 50])

    # ----- All Errors -----
    ws2 = wb.create_sheet("All Errors")
    for i, h in enumerate(SPAN_HEADERS, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(SPAN_HEADERS))
    for ri, s in enumerate(spans, start=2):
        sev = severity_for_family(s["error_family"], analysis)
        for c, fld in enumerate(SPAN_FIELDS, 1):
            cell = ws2.cell(row=ri, column=c, value=s[fld])
            cell.font = MONO if fld in ("span_id", "trace_id", "user_id", "session_id") else BODY
            cell.alignment = LEFT; cell.border = BORDER; cell.fill = fill_for(sev)
    widths(ws2, [14, 22, 36, 30, 30, 34, 10, 8, 30, 34, 22, 30, 32, 70])
    ws2.freeze_panes = "A2"
    if len(spans):
        ws2.auto_filter.ref = ws2.dimensions

    # ----- By Trace -----
    ws3 = wb.create_sheet("By Trace")
    for i, h in enumerate(["Project", "Trace ID", "Span Count", "User ID", "Earliest", "Latest", "Error Families"], 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, 1, 7)
    trows = []
    for tid, d in agg["traces"].items():
        starts = sorted(s[0] for s in d["spans"])
        ends = sorted(s[1] for s in d["spans"])
        trows.append((d["project"], tid, len(d["spans"]), d["user"],
                      starts[0] if starts else "", ends[-1] if ends else "",
                      ", ".join(sorted(d["fams"]))))
    trows.sort(key=lambda x: x[4])
    for ri, row in enumerate(trows, start=2):
        for c, v in enumerate(row, 1):
            cell = ws3.cell(row=ri, column=c, value=v)
            cell.font = MONO if c in (2, 4) else BODY
            cell.alignment = LEFT; cell.border = BORDER
    widths(ws3, [14, 36, 12, 30, 30, 30, 50])
    ws3.freeze_panes = "A2"
    if trows:
        ws3.auto_filter.ref = ws3.dimensions

    # ----- By User -----
    ws4 = wb.create_sheet("By User")
    for i, h in enumerate(["User ID", "Project", "Error Count", "Distinct Traces", "Notes"], 1):
        ws4.cell(row=1, column=i, value=h)
    style_header(ws4, 1, 5)
    notes = analysis.get("user_notes", {})
    urows = []
    for (user, project), d in agg["users"].items():
        urows.append((user, project, d["count"], len(d["traces"]), notes.get(user, "")))
    urows.sort(key=lambda x: -x[2])
    for ri, row in enumerate(urows, start=2):
        for c, v in enumerate(row, 1):
            cell = ws4.cell(row=ri, column=c, value=v)
            cell.font = MONO if c == 1 else BODY
            cell.alignment = LEFT; cell.border = BORDER
    widths(ws4, [34, 16, 14, 16, 50])
    ws4.freeze_panes = "A2"

    # ----- Recommendations -----
    ws5 = wb.create_sheet("Recommendations")
    ws5["A1"] = "Recommended Actions"; ws5["A1"].font = TITLE_FONT; ws5.merge_cells("A1:C1")
    for i, h in enumerate(["Priority", "Issue", "Action"], 1):
        ws5.cell(row=3, column=i, value=h)
    style_header(ws5, 3, 3)
    for ri, rec in enumerate(analysis.get("recommendations", []), start=4):
        sev = rec.get("priority", "P2")
        vals = [sev, rec.get("issue", ""), rec.get("action", "")]
        for c, v in enumerate(vals, 1):
            cell = ws5.cell(row=ri, column=c, value=v)
            cell.font = BODY; cell.alignment = WRAP; cell.border = BORDER
            cell.fill = fill_for(sev)
        ws5.row_dimensions[ri].height = 64
    widths(ws5, [10, 50, 80])

    # ----- Diagnostics (inserted right after Summary) -----
    wsd = wb.create_sheet("Diagnostics", 1)
    r = 1
    wsd.cell(row=r, column=1, value="Diagnostics").font = TITLE_FONT
    wsd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    legend = ("How to read: times are in seconds, tokens are counts. "
              "Average = mean; Median (p50) = typical case (half are faster); "
              "95th pct (p95) = slow tail (only the worst 5% are above it). "
              "Latency is split into Failed (time-to-fail, before the error) and Healthy (successful operations).")
    lc = wsd.cell(row=r, column=1, value=legend)
    lc.font = Font(name="Arial", size=9, italic=True, color="6B7785")
    lc.alignment = WRAP
    wsd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    wsd.row_dimensions[r].height = 42
    r += 2

    def section(title, headers):
        nonlocal r
        wsd.cell(row=r, column=1, value=title).font = SUB_FONT
        r += 1
        for i, h in enumerate(headers, 1):
            wsd.cell(row=r, column=i, value=h)
        style_header(wsd, r, len(headers))
        r += 1

    def row(vals, mono_cols=()):
        nonlocal r
        for c, v in enumerate(vals, 1):
            cell = wsd.cell(row=r, column=c, value=v)
            cell.font = MONO if c in mono_cols else BODY
            cell.alignment = LEFT; cell.border = BORDER
        r += 1

    def fmt(x, suffix=""):
        return "" if x is None else (f"{x:.2f}{suffix}" if isinstance(x, float) else f"{x}{suffix}")

    # Error rate
    section("Error rate by project", ["Project", "Errors", "Total", "Error %", "Basis"])
    for p, d in diag["error_rate"].items():
        row([p, d["errors"], d["total"] if d["total"] is not None else "?",
             fmt(d["pct"], "%") if d["pct"] is not None else "n/a", d["basis"]])
    r += 1

    # Latency
    section("Latency — Failed spans (time-to-fail, seconds)", ["Project", "Span kind", "Count", "Average", "Median (p50)", "95th pct (p95)"])
    for (proj, kind), st in diag["latency_error"].items():
        if st:
            row([proj, kind, st["n"], fmt(st["avg"]), fmt(st["p50"]), fmt(st["p95"])])
    r += 1
    if diag["has_ok"]:
        section("Latency — Healthy spans (successful, seconds)", ["Project", "Span kind", "Count", "Average", "Median (p50)", "95th pct (p95)"])
        for (proj, kind), st in diag["latency_ok"].items():
            if st:
                row([proj, kind, st["n"], fmt(st["avg"]), fmt(st["p50"]), fmt(st["p95"])])
        r += 1

    # Tokens
    section("Token usage by model", ["Model", "Prompt average", "Prompt 95th pct", "Completion average", "Total average", "Total 95th pct"])
    if diag["tokens"]:
        for m, t in diag["tokens"].items():
            pa = t.get("prompt"); ca = t.get("completion"); ta = t.get("total")
            row([m, fmt(pa["avg"]) if pa else "", fmt(pa["p95"]) if pa else "",
                 fmt(ca["avg"]) if ca else "", fmt(ta["avg"]) if ta else "", fmt(ta["p95"]) if ta else ""])
    else:
        row(["(no token counts captured on these spans)"])
    r += 1

    # Concentration + loops
    section(f"Top users by errors (top user = {diag['top_user_pct']:.0f}% of all errors)", ["User ID", "Error count"])
    for u, c in diag["top_users"]:
        row([u, c], mono_cols=(1,))
    r += 1
    section("Redelivery loops (same session + message-id repeated)", ["Session ID", "Message ID (mid)", "Repeats"])
    if diag["loops"]:
        for lp in diag["loops"]:
            row([lp["session"], lp["mid"][:32], lp["count"]], mono_cols=(1, 2))
    else:
        row(["(none detected)"])
    r += 1

    # Errors over time
    section("Errors over time (hourly, UTC)", ["Hour", "Errors", "Bar"])
    mx = max((c for _, c in diag["over_time"]), default=1)
    for h, c in diag["over_time"]:
        row([h.replace("T", " ") + ":00", c, "█" * max(1, round(c / mx * 40))])

    widths(wsd, [40, 20, 16, 16, 16, 16])

    wb.save(out_path)
    return out_path


# ---------------- HTML ----------------

def build_html(spans, analysis, agg, diag, out_path):
    e = html.escape

    def sev_badge(sev):
        color = SEV_COLORS.get(sev, DEFAULT_FILL)
        return f'<span class="badge" style="background:#{color}">{e(sev)}</span>'

    def f2(x, suf=""):
        return "" if x is None else (f"{x:.2f}{suf}" if isinstance(x, float) else f"{x}{suf}")

    # --- Diagnostics rows ---
    rows_rate = ""
    for p, d in diag["error_rate"].items():
        pct = f"{d['pct']:.1f}%" if d["pct"] is not None else "n/a"
        rows_rate += (f"<tr><td>{e(p)}</td><td class='num'>{d['errors']}</td>"
                      f"<td class='num'>{d['total'] if d['total'] is not None else '?'}</td>"
                      f"<td class='num'>{pct}</td><td>{e(d['basis'])}</td></tr>")
    rows_lat = ""
    for label, tbl in (("error", diag["latency_error"]), ("ok", diag["latency_ok"])):
        if label == "ok" and not diag["has_ok"]:
            continue
        tag = "Failed (time-to-fail)" if label == "error" else "Healthy (successful)"
        for (proj, kind), st in tbl.items():
            if st:
                rows_lat += (f"<tr><td>{e(proj)}</td><td>{e(kind)}</td><td>{e(tag)}</td>"
                             f"<td class='num'>{st['n']}</td><td class='num'>{f2(st['avg'])}</td>"
                             f"<td class='num'>{f2(st['p50'])}</td><td class='num'>{f2(st['p95'])}</td></tr>")
    rows_tok = ""
    for m, t in diag["tokens"].items():
        pa, ca, ta = t.get("prompt"), t.get("completion"), t.get("total")
        rows_tok += (f"<tr><td>{e(m)}</td>"
                     f"<td class='num'>{f2(pa['avg']) if pa else ''}</td><td class='num'>{f2(pa['p95']) if pa else ''}</td>"
                     f"<td class='num'>{f2(ca['avg']) if ca else ''}</td>"
                     f"<td class='num'>{f2(ta['avg']) if ta else ''}</td><td class='num'>{f2(ta['p95']) if ta else ''}</td></tr>")
    if not rows_tok:
        rows_tok = "<tr><td colspan=6>No token counts captured on these spans.</td></tr>"
    rows_loop = ""
    for lp in diag["loops"]:
        rows_loop += (f"<tr><td>{e(lp['session'])}</td><td>{e(lp['mid'][:32])}…</td>"
                      f"<td class='num'>{lp['count']}</td></tr>")
    if not rows_loop:
        rows_loop = "<tr><td colspan=3>None detected.</td></tr>"
    mx = max((c for _, c in diag["over_time"]), default=1)
    rows_time = ""
    for h, c in diag["over_time"]:
        bar = "█" * max(1, round(c / mx * 40))
        rows_time += (f"<tr><td>{e(h.replace('T', ' '))}:00</td><td class='num'>{c}</td>"
                      f"<td style='font-family:monospace;color:#1F4E78'>{bar}</td></tr>")

    rows_fam = ""
    for fam in agg["ordered_fams"]:
        cnt = agg["fam_counts"][fam]
        pct = (cnt / agg["total"] * 100) if agg["total"] else 0
        rc, action = family_meta(fam, analysis)
        sev = severity_for_family(fam, analysis)
        rows_fam += (
            f'<tr style="background:#{SEV_COLORS.get(sev, DEFAULT_FILL)}33">'
            f"<td>{e(fam)} {sev_badge(sev)}</td><td class='num'>{cnt}</td>"
            f"<td class='num'>{pct:.1f}%</td><td>{e(rc)}</td><td>{e(action)}</td></tr>"
        )

    rows_proj = ""
    for name, d in agg["projects"].items():
        rows_proj += (
            f"<tr><td>{e(name)}</td><td class='num'>{d['errors']}</td>"
            f"<td class='num'>{len(d['traces'])}</td><td class='num'>{len(d['users'])}</td>"
            f"<td>{e(d.get('note', ''))}</td></tr>"
        )

    rows_rec = ""
    for rec in analysis.get("recommendations", []):
        sev = rec.get("priority", "P2")
        rows_rec += (
            f'<tr style="background:#{SEV_COLORS.get(sev, DEFAULT_FILL)}33">'
            f"<td>{sev_badge(sev)}</td><td>{e(rec.get('issue', ''))}</td>"
            f"<td>{e(rec.get('action', ''))}</td></tr>"
        )

    title = e(analysis.get("report_title", "Phoenix Error Report"))
    meta = e(f"Date: {analysis.get('date', '')}  ·  Window: {analysis.get('window', '')}  ·  Source: {analysis.get('source', '')}")

    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>
  :root {{ --ink:#1f2933; --accent:#1F4E78; --line:#e3e8ee; }}
  * {{ box-sizing:border-box; }}
  body {{ font:15px/1.55 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
    color:var(--ink); max-width:1000px; margin:0 auto; padding:32px 24px 80px; background:#fff; }}
  h1 {{ color:var(--accent); font-size:26px; margin:0 0 4px; }}
  .meta {{ color:#6b7785; font-size:13px; margin-bottom:24px; }}
  h2 {{ color:var(--accent); font-size:18px; margin:32px 0 10px; }}
  table {{ border-collapse:collapse; width:100%; margin:8px 0 4px; font-size:13.5px; }}
  th {{ background:var(--accent); color:#fff; text-align:left; padding:8px 10px; font-weight:600; }}
  td {{ padding:8px 10px; border:1px solid var(--line); vertical-align:top; }}
  td.num {{ text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }}
  .badge {{ display:inline-block; padding:1px 7px; border-radius:10px; font-size:11px;
    font-weight:700; color:#5a3030; margin-left:4px; }}
  .totals {{ display:flex; gap:24px; flex-wrap:wrap; margin:8px 0 4px; }}
  .totals div {{ background:#f5f7fa; border:1px solid var(--line); border-radius:8px; padding:10px 16px; }}
  .totals b {{ display:block; font-size:22px; color:var(--accent); }}
  .totals span {{ font-size:12px; color:#6b7785; }}
  #copy {{ position:fixed; top:16px; right:16px; background:var(--accent); color:#fff; border:0;
    border-radius:8px; padding:9px 16px; font-size:13px; cursor:pointer; box-shadow:0 1px 4px #0002; }}
  #copy:active {{ transform:translateY(1px); }}
</style></head><body>
<button id="copy" onclick="copyReport()">Copy report</button>
<div id="report">
<h1>{title}</h1>
<div class="meta">{meta}</div>

<div class="totals">
  <div><b>{agg['total']}</b><span>error spans</span></div>
  <div><b>{agg['distinct_traces']}</b><span>distinct traces</span></div>
  <div><b>{agg['distinct_users']}</b><span>users affected</span></div>
  <div><b>{e(agg['window'])}</b><span>window (UTC)</span></div>
</div>

<h2>By Project</h2>
<table><thead><tr><th>Project</th><th>Errors</th><th>Traces</th><th>Users</th><th>Note</th></tr></thead>
<tbody>{rows_proj}</tbody></table>

<h2>By Error Family</h2>
<table><thead><tr><th>Error Family</th><th>Count</th><th>% of Total</th><th>Root Cause</th><th>Recommended Action</th></tr></thead>
<tbody>{rows_fam}</tbody></table>

<h2>Recommendations</h2>
<table><thead><tr><th>Priority</th><th>Issue</th><th>Action</th></tr></thead>
<tbody>{rows_rec or '<tr><td colspan=3>No recommendations supplied.</td></tr>'}</tbody></table>

<h2>Diagnostics</h2>
<p style="color:#6b7785;font-size:12.5px;margin:0 0 12px">
  <b>How to read the numbers:</b> times are in seconds and tokens are counts.
  <b>Average</b> = the mean. <b>Median (p50)</b> = the typical case — half are faster, half slower.
  <b>95th percentile (p95)</b> = the slow tail — only the worst 5% are above this.
  Latency rows are split into <b>Failed (time-to-fail)</b> = how long an operation churned before erroring, and
  <b>Healthy (successful)</b> = how long successful operations actually took.
</p>
<h3 style="color:#1F4E78;font-size:15px;margin:18px 0 4px">Error rate by project</h3>
<table><thead><tr><th>Project</th><th>Errors</th><th>Total</th><th>Error %</th><th>Basis</th></tr></thead>
<tbody>{rows_rate}</tbody></table>

<h3 style="color:#1F4E78;font-size:15px;margin:18px 0 4px">Latency (seconds) — by project &amp; span kind</h3>
<table><thead><tr><th>Project</th><th>Span kind</th><th>Span set</th><th>Count</th><th>Average</th><th>Median (p50)</th><th>95th pct (p95)</th></tr></thead>
<tbody>{rows_lat}</tbody></table>

<h3 style="color:#1F4E78;font-size:15px;margin:18px 0 4px">Token usage by model</h3>
<table><thead><tr><th>Model</th><th>Prompt average</th><th>Prompt 95th pct</th><th>Completion average</th><th>Total average</th><th>Total 95th pct</th></tr></thead>
<tbody>{rows_tok}</tbody></table>

<h3 style="color:#1F4E78;font-size:15px;margin:18px 0 4px">Redelivery loops — top user = {diag['top_user_pct']:.0f}% of all errors</h3>
<table><thead><tr><th>Session ID</th><th>Message ID (mid)</th><th>Repeats</th></tr></thead>
<tbody>{rows_loop}</tbody></table>

<h3 style="color:#1F4E78;font-size:15px;margin:18px 0 4px">Errors over time (hourly, UTC)</h3>
<table><thead><tr><th>Hour</th><th>Errors</th><th>Volume</th></tr></thead>
<tbody>{rows_time}</tbody></table>
</div>
<script>
function copyReport() {{
  const r = document.getElementById('report');
  const sel = window.getSelection(); const range = document.createRange();
  range.selectNodeContents(r); sel.removeAllRanges(); sel.addRange(range);
  try {{ document.execCommand('copy'); const b=document.getElementById('copy');
    b.textContent='Copied!'; setTimeout(()=>b.textContent='Copy report',1500); }} catch(e){{}}
  sel.removeAllRanges();
}}
</script>
</body></html>"""
    with open(out_path, "w") as f:
        f.write(doc)
    return out_path


# ---------------- main ----------------

def main():
    ap = argparse.ArgumentParser(description="Render a Phoenix error report (xlsx + html).")
    ap.add_argument("--spans", required=True, help="Path to spans.json (error spans)")
    ap.add_argument("--ok-spans", dest="ok_spans", help="Path to ok_spans.json — healthy span sample (optional; enables healthy latency, error-rate, token trends)")
    ap.add_argument("--analysis", help="Path to analysis.json (optional)")
    ap.add_argument("--out", required=True, help="Output path prefix (no extension)")
    ap.add_argument("--format", choices=["both", "xlsx", "html"], default="both")
    args = ap.parse_args()

    spans = load_spans(args.spans)
    ok_spans = load_spans(args.ok_spans, default_status="OK") if args.ok_spans else []
    analysis = load_analysis(args.analysis)
    agg = aggregate(spans, analysis)
    diag = diagnostics(spans, ok_spans, analysis)

    written = []
    if args.format in ("both", "xlsx"):
        try:
            written.append(build_xlsx(spans, analysis, agg, diag, args.out + ".xlsx"))
        except ImportError:
            print("openpyxl not installed — skipping xlsx. `pip install openpyxl` to enable.", file=sys.stderr)
    if args.format in ("both", "html"):
        written.append(build_html(spans, analysis, agg, diag, args.out + ".html"))

    print(f"Rendered {agg['total']} error spans across {len(agg['projects'])} project(s).")
    for w in written:
        print(f"  -> {w}")


if __name__ == "__main__":
    main()
